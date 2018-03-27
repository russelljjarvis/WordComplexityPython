#call in necessary packages that code requires to operate and define additional vars
import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import re

def trim(test_string, removal_string):
    return re.sub(r'^(.*?)('+ removal_string + ')(.*)$', r'\1' + r'\2', test_string)
##################################################
##################################################

#define number of loop iterations (i.e. how many links to analyze)
num = 26 #how many links + 1
term = 'Transgenic'

#load working directory for file saving of files containing the URL text
os.chdir('/Users/PMcG/Dropbox (ASU)/AAB_files/Pat-files/WCP/code/Data Files/'+ str(term) +'/')

#load in the excel workbook
wb = openpyxl.load_workbook('xml_searchResults_' + str(term) + '.xlsx')

#retrieve the sheet names
wb.get_sheet_names()

#use the called names to define objects for each search engine
google = wb.get_sheet_by_name('google')
bing = wb.get_sheet_by_name('bing') 
yahoo = wb.get_sheet_by_name('yahoo')

##################################################
##################################################
                                    ###GOOGLE###
##################################################
##################################################

#define parameters for url text removal
prefix1 = ("/url?q=")
prefix2 = ("&sa=")

for x in range (1,num):

    #remove excess text from each URL
    temp_dat = str(google.cell(row = x, column = 1).value)   
    urlReady = temp_dat.replace(prefix1, "", 1)
    urlReady = trim(urlReady, '&sa=')
    urlReady = urlReady.replace(prefix2, "", 1)

    #verify URL is correct and ready to be accessed via web server
    print x
    print urlReady
    
    google.cell(row = x, column = 2).value = urlReady

##################################################
##################################################
                                    ###YAHOO###
##################################################
##################################################

#define parameters for url text removal
prefix1 = ('/RK=')

for x in range (1,num):

    #remove excess text from each URL
    temp_dat = str(yahoo.cell(row = x, column = 1).value)
    urlReady = trim(temp_dat, '/RK=')
    urlReady = urlReady.split("RO=10/RU=").pop()
    urlReady = urlReady.replace("%2f", "/")
    urlReady = urlReady.replace("%3a", ":")
    urlReady = urlReady.replace(prefix1, "", 1)

    #verify URL is correct and ready to be accessed via web server
    print x
    print urlReady

    yahoo.cell(row = x, column = 2).value = urlReady
       
##################################################
##################################################
                                    ###BING###
##################################################
##################################################

for x in range (1,num):

    #remove excess text from each URL
    urlReady = str(bing.cell(row = x, column = 1).value)   

    #verify URL is correct and ready to be accessed via web server
    print x
    print urlReady

    bing.cell(row = x, column = 2).value = urlReady
       
wb.save('xml_searchResults_' + str(term) + '.xlsx')
